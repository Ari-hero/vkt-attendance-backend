import json
import logging
import uuid as uuid_lib
from pathlib import Path
from datetime import datetime
from django.utils import timezone
from django.http import HttpResponse
from rest_framework import status, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenPyXLImage

from .models import Device, Employee, AttendanceLog
from .serializers import DeviceSerializer, EmployeeSerializer, AttendanceLogSerializer

logger = logging.getLogger('api')


def authenticate_device(request):
    """
    Validates X-Device-Id and X-Api-Key headers.
    Returns (device_instance, None) on success or (None, response_tuple) on failure.
    Updates device.last_seen on every successful call.
    """
    device_id = request.headers.get('X-Device-Id')
    api_key = request.headers.get('X-Api-Key')

    if not device_id or not api_key:
        logger.warning(f"Auth failed: Missing device headers (Device: {device_id})")
        return None, (Response({'error': 'Missing authentication headers (X-Device-Id, X-Api-Key)'}, status=status.HTTP_401_UNAUTHORIZED))

    try:
        device = Device.objects.get(device_id=device_id)
        if device.api_key != api_key:
            logger.warning(f"Auth failed: Invalid API key for device {device_id}")
            return None, (Response({'error': 'Invalid API key'}, status=status.HTTP_401_UNAUTHORIZED))

        if not device.is_active:
            logger.warning(f"Auth failed: Inactive device attempted access {device_id}")
            return None, (Response({'error': 'Device account is disabled'}, status=status.HTTP_403_FORBIDDEN))

        # Update health monitoring timestamp
        Device.objects.filter(id=device.id).update(last_seen=timezone.now())
        return device, None
    except Device.DoesNotExist:
        logger.warning(f"Auth failed: Unregistered device {device_id}")
        return None, (Response({'error': 'Unregistered device ID'}, status=status.HTTP_401_UNAUTHORIZED))


class RegisterDeviceView(APIView):
    def post(self, request):
        device_id = request.data.get('device_id')
        name = request.data.get('device_name', 'Industrial Kiosk')
        if not device_id:
            device_id = f"kiosk_{uuid_lib.uuid4().hex[:8]}"

        device, created = Device.objects.get_or_create(
            device_id=device_id,
            defaults={'name': name, 'api_key': str(uuid_lib.uuid4()), 'last_seen': timezone.now()}
        )
        if not created:
            device.last_seen = timezone.now()
            device.save()

        logger.info(f"Device registered/updated: {device_id} ({name})")
        return Response(DeviceSerializer(device).data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


def is_authenticated_request(request):
    """
    Validates if request is authorized via DRF Token/Session (web dashboard)
    or via valid X-Device-Id / X-Api-Key headers (kiosk terminal).
    """
    if request.user and request.user.is_authenticated:
        return True

    device_id = request.headers.get('X-Device-Id')
    api_key = request.headers.get('X-Api-Key')
    if device_id and api_key:
        try:
            device = Device.objects.get(device_id=device_id, is_active=True)
            if device.api_key == api_key:
                Device.objects.filter(id=device.id).update(last_seen=timezone.now())
                return True
        except Device.DoesNotExist:
            pass
    return False


class DeviceListView(generics.ListAPIView):
    queryset = Device.objects.all().order_by('-last_seen')
    serializer_class = DeviceSerializer

    def list(self, request, *args, **kwargs):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)
        return super().list(request, *args, **kwargs)


class StandardResultsPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 500


class EmployeeListCreateView(generics.ListCreateAPIView):
    queryset = Employee.objects.all().order_by('name')
    serializer_class = EmployeeSerializer

    def list(self, request, *args, **kwargs):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)
        return super().list(request, *args, **kwargs)



class SyncAttendanceView(APIView):
    """
    Accepts raw attendance logs from kiosks.
    Performs device auth, payload validation, deduplication by UUID, and global IN/OUT sequence resolution.
    Returns backend-assigned final_type for each log.
    """
    def post(self, request):
        device, err_response = authenticate_device(request)
        if err_response:
            return err_response

        logs_data = request.data
        if not isinstance(logs_data, list):
            logs_data = [logs_data]

        # Security check: payload size limit
        if len(logs_data) > 500:
            logger.warning(f"Device {device.device_id} sent batch exceeding 500 limit ({len(logs_data)})")
            return Response({'error': 'Batch size exceeds 500 logs limit'}, status=status.HTTP_400_BAD_REQUEST)

        processed_list = []
        incoming_uuids = [item.get('uuid') for item in logs_data if item.get('uuid')]

        # Query existing UUIDs in bulk for performance
        existing_logs = {
            log.uuid: log for log in AttendanceLog.objects.filter(uuid__in=incoming_uuids)
        }

        new_logs_to_create = []

        for item in logs_data:
            uuid_str = item.get('uuid')
            emp_id = item.get('emp_id')
            emp_name = item.get('emp_name')
            timestamp_str = item.get('timestamp')
            date_str = item.get('date')
            time_str = item.get('time')
            confidence = item.get('confidence', 1.0)

            if not uuid_str or not emp_id:
                continue

            # Idempotency check: if already stored, return existing backend classification
            if uuid_str in existing_logs:
                existing = existing_logs[uuid_str]
                processed_list.append({
                    'uuid': uuid_str,
                    'emp_id': emp_id,
                    'final_type': existing.type,
                    'server_timestamp': existing.synced_at.isoformat(),
                })
                continue

            try:
                dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            except Exception:
                dt = timezone.now()

            # Global IN/OUT Sequence Resolution across ALL devices for this employee
            last_log = AttendanceLog.objects.filter(emp_id=emp_id).order_by('-timestamp').first()
            assigned_type = 'OUT' if (last_log and last_log.type == 'IN') else 'IN'

            new_log = AttendanceLog(
                uuid=uuid_str,
                emp_id=emp_id,
                emp_name=emp_name or f"Emp {emp_id}",
                timestamp=dt,
                date=date_str or dt.strftime('%Y-%m-%d'),
                time=time_str or dt.strftime('%H:%M:%S'),
                type=assigned_type,
                confidence=confidence,
                device_id=device.device_id,
            )
            new_logs_to_create.append(new_log)
            processed_list.append({
                'uuid': uuid_str,
                'emp_id': emp_id,
                'final_type': assigned_type,
                'server_timestamp': timezone.now().isoformat(),
            })

        # Bulk insert new logs safely
        if new_logs_to_create:
            AttendanceLog.objects.bulk_create(new_logs_to_create, ignore_conflicts=True)
            logger.info(f"Device {device.device_id} synced {len(new_logs_to_create)} new logs.")

        return Response({
            'status': 'success',
            'synced_count': len(processed_list),
            'processed': processed_list,
        }, status=status.HTTP_200_OK)


class AttendanceLogListView(generics.ListAPIView):
    serializer_class = AttendanceLogSerializer
    pagination_class = StandardResultsPagination

    def get_queryset(self):
        queryset = AttendanceLog.objects.all().order_by('-timestamp')
        date_param = self.request.query_params.get('date')
        emp_id = self.request.query_params.get('emp_id')
        log_type = self.request.query_params.get('type')

        if date_param:
            queryset = queryset.filter(date=date_param)
        if emp_id:
            queryset = queryset.filter(emp_id=emp_id)
        if log_type and log_type != 'ALL':
            queryset = queryset.filter(type=log_type)

        return queryset

    def list(self, request, *args, **kwargs):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)
        return super().list(request, *args, **kwargs)


class ExportExcelReportView(APIView):
    """
    Generates detailed Excel report with embedded corporate logo graphics,
    corporate header hierarchy, alternating row shading, and individual event rows.
    """
    def get(self, request):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)
        target_date_str = request.query_params.get('date', datetime.now().strftime('%Y-%m-%d'))

        
        all_employees = Employee.objects.all().order_by('name')
        day_logs = AttendanceLog.objects.filter(date=target_date_str).order_by('timestamp')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Report {target_date_str}"

        # ── Embed Corporate Logo Graphics ─────────────────────────────────────
        base_dir = Path(__file__).resolve().parent.parent.parent
        sp_logo_path = base_dir / 'assets' / 'images' / 'logo_spinfotech.jpeg'
        vkt_logo_path = base_dir / 'assets' / 'images' / 'logo_vkt.jpg'

        if sp_logo_path.exists():
            try:
                img_sp = OpenPyXLImage(str(sp_logo_path))
                img_sp.width = 135
                img_sp.height = 48
                ws.add_image(img_sp, 'A1')
            except Exception as e:
                logger.warning(f"Could not embed S.P.Infotech logo: {e}")

        if vkt_logo_path.exists():
            try:
                img_vkt = OpenPyXLImage(str(vkt_logo_path))
                img_vkt.width = 115
                img_vkt.height = 42
                ws.add_image(img_vkt, 'G1')
            except Exception as e:
                logger.warning(f"Could not embed VKT logo: {e}")

        # Row Heights & Spacing
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 12
        ws.row_dimensions[5].height = 26

        # Corporate Header Text Hierarchy
        c1 = ws.cell(row=1, column=3, value="PRIMARY SOFTWARE PROVIDER: S.P. INFOTECH")
        c1.font = Font(size=11, bold=True, color="1E293B")
        c2 = ws.cell(row=2, column=3, value="CLIENT ENTERPRISE: V.K. TOURS & TRAVELS")
        c2.font = Font(size=10, bold=False, color="475569")
        c3 = ws.cell(row=3, column=3, value=f"DAILY DETAILED ATTENDANCE REPORT — DATE: {target_date_str}")
        c3.font = Font(size=10, bold=True, color="1D4ED8")

        headers = ['Employee ID', 'Employee Name', 'Department', 'Date', 'Time', 'Punch Type', 'Confidence Score', 'Timestamp', 'Record UUID']
        ws.append([]) # row 4 blank

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        header_row_idx = 5
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row_idx, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        logs_by_emp = {}
        for log in day_logs:
            logs_by_emp.setdefault(log.emp_id, []).append(log)

        all_emp_ids = list(dict.fromkeys([e.emp_id for e in all_employees] + list(logs_by_emp.keys())))

        even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        current_row_idx = 6
        for emp_id in all_emp_ids:
            emp = next((e for e in all_employees if e.emp_id == emp_id), None)
            emp_name = emp.name if emp else (logs_by_emp[emp_id][0].emp_name if logs_by_emp.get(emp_id) else emp_id)
            department = emp.department if emp else 'Unregistered'
            emp_logs = logs_by_emp.get(emp_id, [])

            rows_to_add = []
            if not emp_logs:
                rows_to_add.append([emp_id, emp_name, department, target_date_str, '-', 'ABSENT', '-', '-', '-'])
            else:
                for log in emp_logs:
                    rows_to_add.append([
                        log.emp_id,
                        log.emp_name,
                        department,
                        log.date,
                        log.time.strftime('%H:%M:%S') if hasattr(log.time, 'strftime') else str(log.time),
                        log.type,
                        f"{log.confidence:.3f}",
                        log.timestamp.isoformat() if hasattr(log.timestamp, 'isoformat') else str(log.timestamp),
                        log.uuid
                    ])

            for row_values in rows_to_add:
                row_fill = even_fill if current_row_idx % 2 == 0 else odd_fill
                for col_idx, val in enumerate(row_values, 1):
                    c = ws.cell(row=current_row_idx, column=col_idx, value=val)
                    c.fill = row_fill
                    c.border = thin_border
                    c.font = Font(size=9, color="0F172A")
                    if col_idx in (1, 4, 5, 6, 7):
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[current_row_idx].height = 20
                current_row_idx += 1

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.row >= 5 and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Attendance_Report_{target_date_str}.xlsx"'
        wb.save(response)
        return response
