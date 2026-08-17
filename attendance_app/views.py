import os
import json
import logging
import uuid as uuid_lib
from pathlib import Path
from datetime import datetime
from django.db import models
from django.db.models import F
from django.utils import timezone
from django.http import HttpResponse
from django.conf import settings
from rest_framework import status, generics
from rest_framework.views import APIView
from rest_framework.response import Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenPyXLImage

from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from rest_framework.permissions import AllowAny

from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors as rl_colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import Device, Employee, AttendanceLog
from .serializers import DeviceSerializer, EmployeeSerializer, AttendanceLogSerializer

logger = logging.getLogger('api')


def is_authenticated_request(request):
    """
    Validates if request is authorized via DRF Token/Session (web dashboard)
    or via valid X-Device-Id / X-Api-Key headers (kiosk terminal).
    """
    if request.user and request.user.is_authenticated:
        return True

    device_id = request.headers.get('X-Device-Id')
    api_key = request.headers.get('X-Api-Key')
    if device_id and api_key:
        try:
            device = Device.objects.get(device_id=device_id, is_active=True)
            if device.api_key == api_key:
                Device.objects.filter(id=device.id).update(last_seen=timezone.now())
                return True
        except Device.DoesNotExist:
            pass
    return False


class ApiRootView(APIView):
    """
    Root API Endpoint returning online status and service details.
    """
    permission_classes = [AllowAny]

    def get(self, request):
        return Response({
            "status": "online",
            "service": "VKT Biometric Attendance API",
            "version": "1.0.0",
            "endpoints": {
                "admin_login": "/api/auth/login/",
                "provision_otp": "/api/devices/provision-otp/",
                "register_device": "/api/register-device/",
                "employees": "/api/employees/",
                "sync_attendance": "/api/sync-attendance/",
                "devices": "/api/devices/",
                "attendance": "/api/attendance/",
                "daily_data": "/api/reports/daily-data/",
                "export_excel": "/api/reports/export-excel/",
                "export_pdf": "/api/reports/export-pdf/",
            }
        }, status=status.HTTP_200_OK)


class AdminLoginView(APIView):
    """
    DRF Token Login endpoint for Dashboard Administrators.
    Validates username and password against Django auth_user database.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        if not username or not password:
            return Response({'error': 'Username and password are required'}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(request, username=username, password=password)
        if user is not None:
            if not user.is_active:
                return Response({'error': 'User account is disabled'}, status=status.HTTP_403_FORBIDDEN)
            token, _ = Token.objects.get_or_create(user=user)
            return Response({
                'token': token.key,
                'username': user.username,
                'is_staff': user.is_staff
            }, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Invalid username or password'}, status=status.HTTP_401_UNAUTHORIZED)


import hashlib
import secrets
from .models import Device, Employee, AttendanceLog, DeviceProvisioningOTP
from .serializers import DeviceSerializer, EmployeeSerializer, AttendanceLogSerializer


class GenerateProvisioningOTPView(APIView):
    """
    Admin-only endpoint to generate a short-lived (10 min), single-use 6-digit OTP
    for pairing a new hardware kiosk without exposing admin credentials or API keys.
    """
    def post(self, request):
        if not (request.user and request.user.is_authenticated):
            return Response({'error': 'Administrator authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

        device_name = request.data.get('device_name', 'Industrial Kiosk')
        duration_minutes = int(request.data.get('duration_minutes', 10))

        # Rate limiting: max 5 active OTPs created in last 10 minutes
        ten_mins_ago = timezone.now() - timezone.timedelta(minutes=10)
        recent_count = DeviceProvisioningOTP.objects.filter(
            created_at__gte=ten_mins_ago,
            is_active=True
        ).count()
        if recent_count >= 5:
            return Response({'error': 'Rate limit exceeded. Too many active OTP requests. Please wait or use existing code.'}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        # Deactivate any previous unused OTPs for the same device name
        DeviceProvisioningOTP.objects.filter(
            device_name=device_name,
            is_active=True
        ).update(is_active=False)

        # Generate 6-digit numeric OTP
        raw_otp = f"{secrets.randbelow(900000) + 100000}"
        otp_hash = hashlib.sha256(raw_otp.encode('utf-8')).hexdigest()
        expires_at = timezone.now() + timezone.timedelta(minutes=duration_minutes)

        otp_obj = DeviceProvisioningOTP.objects.create(
            otp_hash=otp_hash,
            expires_at=expires_at,
            created_by=request.user,
            device_name=device_name,
            is_active=True
        )

        logger.info(f"Provisioning OTP generated for device '{device_name}' by user '{request.user.username}'")

        return Response({
            'otp': raw_otp,
            'device_name': device_name,
            'expires_at': expires_at.isoformat(),
            'expires_in_seconds': duration_minutes * 60,
            'status': 'waiting_for_activation'
        }, status=status.HTTP_201_CREATED)


class RegisterDeviceView(APIView):
    """
    Public device registration endpoint.
    Requires a valid, unexpired, single-use 6-digit OTP generated by an administrator.
    Returns device_id and permanent device api_key on success.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        otp_input = str(request.data.get('otp', '')).strip()
        device_name = request.data.get('device_name', request.data.get('name', 'Industrial Kiosk'))

        if otp_input:
            otp_hash = hashlib.sha256(otp_input.encode('utf-8')).hexdigest()
            now = timezone.now()

            # Find matching active OTP
            try:
                otp_obj = DeviceProvisioningOTP.objects.get(otp_hash=otp_hash, is_active=True)
            except DeviceProvisioningOTP.DoesNotExist:
                # Rate limit defense: increment attempt count on recent active OTPs
                DeviceProvisioningOTP.objects.filter(is_active=True, expires_at__gte=now).update(
                    attempt_count=models.F('attempt_count') + 1
                )
                # Auto-deactivate OTPs exceeding 5 attempts
                DeviceProvisioningOTP.objects.filter(attempt_count__gte=5).update(is_active=False)
                return Response({'error': 'Invalid or expired activation code'}, status=status.HTTP_400_BAD_REQUEST)

            # Check expiration
            if now > otp_obj.expires_at:
                otp_obj.is_active = False
                otp_obj.save()
                return Response({'error': 'Activation code has expired. Please generate a new code from dashboard.'}, status=status.HTTP_400_BAD_REQUEST)

            # Check if already used
            if otp_obj.used_at is not None:
                otp_obj.is_active = False
                otp_obj.save()
                return Response({'error': 'Activation code has already been used'}, status=status.HTTP_400_BAD_REQUEST)

            # Check attempt limit
            if otp_obj.attempt_count >= 5:
                otp_obj.is_active = False
                otp_obj.save()
                return Response({'error': 'Maximum verification attempts exceeded. Code revoked.'}, status=status.HTTP_400_BAD_REQUEST)

            # Mark OTP as consumed (single-use guarantee)
            otp_obj.used_at = now
            otp_obj.is_active = False
            otp_obj.save()

            # Create provisioned Device with unique credentials
            device_id = f"kiosk_{uuid_lib.uuid4().hex[:8]}"
            api_key = str(uuid_lib.uuid4())
            device = Device.objects.create(
                device_id=device_id,
                name=otp_obj.device_name or device_name,
                api_key=api_key,
                is_active=True,
                last_seen=now
            )

            logger.info(f"Device successfully provisioned with OTP: {device.device_id} ({device.name})")

            return Response({
                'status': 'success',
                'device_id': device.device_id,
                'api_key': device.api_key,
                'name': device.name,
                'message': 'Device activated successfully'
            }, status=status.HTTP_201_CREATED)

        # Legacy fallback for backward compatibility
        device_id = request.data.get('device_id')
        if not device_id:
            return Response({'error': 'Activation code (otp) is required for device registration'}, status=status.HTTP_400_BAD_REQUEST)

        device, created = Device.objects.get_or_create(
            device_id=device_id,
            defaults={'name': device_name, 'api_key': str(uuid_lib.uuid4()), 'last_seen': timezone.now()}
        )
        if not created:
            device.last_seen = timezone.now()
            if device_name and device_name != device.name:
                device.name = device_name
            device.save()

        logger.info(f"Device registered/updated: {device_id} ({device_name})")
        return Response(DeviceSerializer(device).data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


class DeviceListView(generics.ListAPIView):
    """Returns only active provisioned devices. Inactive/decommissioned devices are excluded."""
    queryset = Device.objects.filter(is_active=True).order_by('-last_seen')
    serializer_class = DeviceSerializer

    def list(self, request, *args, **kwargs):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class EmployeeListCreateView(generics.ListCreateAPIView):
    queryset = Employee.objects.all().order_by('name')
    serializer_class = EmployeeSerializer

    def list(self, request, *args, **kwargs):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



class SyncAttendanceView(APIView):
    """
    Accepts raw attendance logs from kiosks.
    Performs device auth, payload validation, deduplication by UUID, and global IN/OUT sequence resolution.
    Returns backend-assigned final_type for each log.
    """
    def post(self, request):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid (X-Device-Id, X-Api-Key)'}, status=status.HTTP_401_UNAUTHORIZED)

        device_id = request.headers.get('X-Device-Id')
        if device_id:
            Device.objects.filter(device_id=device_id).update(last_seen=timezone.now())

        logs_data = request.data
        if not isinstance(logs_data, list):
            logs_data = [logs_data]

        if len(logs_data) > 500:
            return Response({'error': 'Batch size exceeds 500 logs limit'}, status=status.HTTP_400_BAD_REQUEST)

        processed_list = []
        incoming_uuids = [item.get('uuid') for item in logs_data if item.get('uuid')]

        existing_logs = {
            log.uuid: log for log in AttendanceLog.objects.filter(uuid__in=incoming_uuids)
        }

        new_logs_to_create = []

        for item in logs_data:
            uuid_str = item.get('uuid')
            emp_id = item.get('emp_id')
            emp_name = item.get('emp_name')
            timestamp_str = item.get('timestamp')
            date_str = item.get('date')
            time_str = item.get('time')
            confidence = item.get('confidence', 1.0)
            dev_id = item.get('device_id', device_id or 'default_kiosk')

            if not uuid_str or not emp_id:
                continue

            if uuid_str in existing_logs:
                existing = existing_logs[uuid_str]
                processed_list.append({
                    'uuid': uuid_str,
                    'emp_id': emp_id,
                    'final_type': existing.type,
                    'server_timestamp': existing.synced_at.isoformat(),
                })
                continue

            try:
                dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            except Exception:
                dt = timezone.now()

            # Global IN/OUT Sequence Resolution across ALL devices for this employee
            last_log = AttendanceLog.objects.filter(emp_id=emp_id).order_by('-timestamp').first()
            assigned_type = 'OUT' if (last_log and last_log.type == 'IN') else 'IN'

            new_log = AttendanceLog(
                uuid=uuid_str,
                emp_id=emp_id,
                emp_name=emp_name or f"Emp {emp_id}",
                timestamp=dt,
                date=date_str or dt.strftime('%Y-%m-%d'),
                time=time_str or dt.strftime('%H:%M:%S'),
                type=assigned_type,
                confidence=confidence,
                device_id=dev_id,
            )
            new_logs_to_create.append(new_log)
            processed_list.append({
                'uuid': uuid_str,
                'emp_id': emp_id,
                'final_type': assigned_type,
                'server_timestamp': timezone.now().isoformat(),
            })

        if new_logs_to_create:
            AttendanceLog.objects.bulk_create(new_logs_to_create, ignore_conflicts=True)
            logger.info(f"Synced {len(new_logs_to_create)} new logs.")

        # Ensure employees enrolled on kiosk are represented in central Employee master
        incoming_emp_data = {}
        for item in logs_data:
            e_id = str(item.get('emp_id', '')).strip()
            e_name = item.get('emp_name', '')
            if e_id and e_id not in incoming_emp_data:
                incoming_emp_data[e_id] = e_name

        if incoming_emp_data:
            existing_emp_ids = set(
                Employee.objects.filter(emp_id__in=incoming_emp_data.keys()).values_list('emp_id', flat=True)
            )
            new_emps_to_create = [
                Employee(
                    emp_id=eid,
                    name=incoming_emp_data[eid] or f"Emp {eid}",
                    department="General",
                    embedding="[]",
                    photo_url="",
                )
                for eid in incoming_emp_data
                if eid not in existing_emp_ids
            ]
            if new_emps_to_create:
                Employee.objects.bulk_create(new_emps_to_create, ignore_conflicts=True)

        return Response({
            'status': 'success',
            'synced_count': len(processed_list),
            'processed': processed_list,
        }, status=status.HTTP_200_OK)


class AttendanceLogListView(generics.ListAPIView):
    serializer_class = AttendanceLogSerializer

    def get_queryset(self):
        queryset = AttendanceLog.objects.all().order_by('-timestamp')
        start_date = self.request.query_params.get('start_date') or self.request.query_params.get('date')
        end_date = self.request.query_params.get('end_date') or self.request.query_params.get('date')
        emp_id = self.request.query_params.get('emp_id')
        emp_ids = self.request.query_params.get('emp_ids')
        log_type = self.request.query_params.get('type')

        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)

        if emp_ids and emp_ids != 'ALL':
            id_list = [e.strip() for e in emp_ids.split(',') if e.strip()]
            if id_list:
                queryset = queryset.filter(emp_id__in=id_list)
        elif emp_id and emp_id != 'ALL':
            queryset = queryset.filter(emp_id=emp_id)

        if log_type and log_type != 'ALL':
            queryset = queryset.filter(type=log_type)

        return queryset

    def list(self, request, *args, **kwargs):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

AttendanceListView = AttendanceLogListView


def format_human_date(date_str):
    """Converts 'YYYY-MM-DD' to 'DD Mon YYYY' (e.g. '10 Jun 2026')"""
    try:
        dt = datetime.strptime(str(date_str).strip(), '%Y-%m-%d')
        return dt.strftime('%d %b %Y')
    except Exception:
        return str(date_str)


def format_dmy_date(date_str):
    """Converts 'YYYY-MM-DD' to 'DD-MM-YYYY'"""
    try:
        dt = datetime.strptime(str(date_str).strip(), '%Y-%m-%d')
        return dt.strftime('%d-%m-%Y')
    except Exception:
        return str(date_str)


def get_canonical_attendance_data(start_date_str=None, end_date_str=None, emp_ids=None):
    """
    Single Canonical Backend Attendance Dataset generator.
    Supports inclusive Date Range (start_date to end_date) and Employee Filtering.
    Guarantees identical data structure, ordering, and event counts for Web Dashboard, PDF, and Excel exports.
    """
    # 1. Normalize dates
    today_str = datetime.now().strftime('%Y-%m-%d')
    if not start_date_str and not end_date_str:
        start_date_str = today_str
        end_date_str = today_str
    elif start_date_str and not end_date_str:
        end_date_str = start_date_str
    elif end_date_str and not start_date_str:
        start_date_str = end_date_str

    if start_date_str > end_date_str:
        start_date_str, end_date_str = end_date_str, start_date_str

    if start_date_str == end_date_str:
        display_date_range = format_human_date(start_date_str)
        file_date_suffix = format_dmy_date(start_date_str)
    else:
        display_date_range = f"{format_human_date(start_date_str)} – {format_human_date(end_date_str)}"
        file_date_suffix = f"{format_dmy_date(start_date_str)}_to_{format_dmy_date(end_date_str)}"

    # 2. Parse & normalize employee filter
    emp_ids_list = None
    if emp_ids and emp_ids != 'ALL':
        if isinstance(emp_ids, list):
            emp_ids_list = [str(e).strip() for e in emp_ids if str(e).strip()]
        else:
            emp_ids_list = [e.strip() for e in str(emp_ids).split(',') if e.strip()]

    # 3. Query employees & logs
    all_employees_qs = Employee.objects.all().order_by('name')
    if emp_ids_list:
        all_employees_qs = all_employees_qs.filter(emp_id__in=emp_ids_list)
    all_employees = list(all_employees_qs)

    logs_qs = AttendanceLog.objects.filter(
        date__gte=start_date_str,
        date__lte=end_date_str
    ).order_by('timestamp')

    if emp_ids_list:
        logs_qs = logs_qs.filter(emp_id__in=emp_ids_list)

    logs = list(logs_qs)

    # 4. Group logs by employee ID
    logs_by_emp = {}
    for log in logs:
        logs_by_emp.setdefault(log.emp_id, []).append(log)

    all_emp_ids = list(dict.fromkeys([e.emp_id for e in all_employees] + list(logs_by_emp.keys())))

    # 5. Formulate Employee Scope label
    if not emp_ids_list:
        emp_filter_label = "All Employees"
    else:
        names = [e.name for e in all_employees]
        if names:
            emp_filter_label = ", ".join(names[:3]) + (f" (+{len(names)-3} more)" if len(names) > 3 else "")
        else:
            emp_filter_label = ", ".join(emp_ids_list)

    # 6. Calculate per-employee presence breakdown and detailed punch records
    employee_summaries = []
    for emp_id in all_emp_ids:
        emp = next((e for e in all_employees if e.emp_id == emp_id), None)
        emp_name = emp.name if emp else (logs_by_emp[emp_id][0].emp_name if logs_by_emp.get(emp_id) else emp_id)
        department = emp.department if emp else 'Unregistered'
        emp_logs = logs_by_emp.get(emp_id, [])

        distinct_dates = sorted(list(set(str(l.date) for l in emp_logs)))
        present_days_count = len(distinct_dates)

        employee_summaries.append({
            'emp_id': emp_id,
            'emp_name': emp_name,
            'department': department,
            'present_days': present_days_count,
            'total_punches': len(emp_logs),
            'dates_present': [format_dmy_date(d) for d in distinct_dates],
            'has_attendance': present_days_count > 0,
        })

    # Detailed punch logs: contains ONLY actual recorded punch events in the date range
    records = []
    for log in logs:
        emp = next((e for e in all_employees if e.emp_id == log.emp_id), None)
        department = emp.department if emp else 'Unregistered'
        display_log_date = format_dmy_date(log.date)
        time_str = log.time.strftime('%H:%M:%S') if hasattr(log.time, 'strftime') else str(log.time)
        ts_str = log.timestamp.strftime('%d-%m-%Y %H:%M:%S') if hasattr(log.timestamp, 'strftime') else str(log.timestamp)

        records.append({
            'emp_id': log.emp_id,
            'emp_name': log.emp_name,
            'department': department,
            'date': display_log_date,
            'raw_date': str(log.date),
            'time': time_str,
            'type': log.type,
            'confidence': f"{log.confidence:.3f}",
            'timestamp': ts_str,
            'uuid': log.uuid,
            'is_present': True,
        })

    present_emp_count = len([s for s in employee_summaries if s['has_attendance']])

    return {
        'start_date': start_date_str,
        'end_date': end_date_str,
        'display_date': display_date_range,
        'display_date_range': display_date_range,
        'file_date_suffix': file_date_suffix,
        'employee_filter_label': emp_filter_label,
        'summary': {
            'start_date': start_date_str,
            'end_date': end_date_str,
            'display_date_range': display_date_range,
            'employee_filter_label': emp_filter_label,
            'total_events': len(records),
            'present_count': present_emp_count,
            'enrolled_count': len(all_emp_ids),
            'total_rows': len(records),
        },
        'employee_summaries': employee_summaries,
        'records': records,
    }


class CanonicalReportDataView(APIView):
    """
    Returns canonical JSON report dataset for audit and PDF/Excel dataset equality verification.
    """
    def get(self, request):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)
        start_date = request.query_params.get('start_date') or request.query_params.get('date')
        end_date = request.query_params.get('end_date') or request.query_params.get('date')
        emp_ids = request.query_params.get('emp_ids') or request.query_params.get('emp_id')

        data = get_canonical_attendance_data(start_date_str=start_date, end_date_str=end_date, emp_ids=emp_ids)
        return Response(data, status=status.HTTP_200_OK)


def get_logo_path(filename):
    """
    Resolves the absolute path to a logo asset across local monorepo, backend repo, and deployment containers.
    """
    base_file = Path(__file__).resolve()
    candidates = [
        base_file.parent.parent / 'assets' / 'images' / filename,
        base_file.parent.parent.parent / 'assets' / 'images' / filename,
        Path(os.getcwd()) / 'assets' / 'images' / filename,
        Path(os.getcwd()) / 'backend' / 'assets' / 'images' / filename,
        base_file.parent.parent / 'dashboard' / 'src' / 'assets' / filename,
        base_file.parent.parent.parent / 'dashboard' / 'src' / 'assets' / filename,
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return None


class ExportExcelReportView(APIView):
    """
    Generates detailed Excel report matching the visual layout of the PDF report:
    - Corporate header hierarchy with embedded logo graphics (S.P. Infotech at A1, VKT at G1)
    - Date range representation & Employee scope badge
    - Executive Summary KPI metrics cards (Total Enrolled, Present Staff, Total Punch Events)
    - Section 1: Employee Presence Summary table (Days Present, Total Punches, Actual Attendance Dates)
    - Section 2: Detailed Punch Logs table (Actual IN/OUT punches without Confidence Score and Record UUID)
    - Totals / Summary footer row
    - Print setup: Landscape A4, fit to page width, repeating header setup.
    """
    def get(self, request):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)

        start_date = request.query_params.get('start_date') or request.query_params.get('date')
        end_date = request.query_params.get('end_date') or request.query_params.get('date')
        emp_ids = request.query_params.get('emp_ids') or request.query_params.get('emp_id')

        report_data = get_canonical_attendance_data(start_date_str=start_date, end_date_str=end_date, emp_ids=emp_ids)
        display_date_range = report_data['display_date_range']
        file_date_suffix = report_data['file_date_suffix']
        emp_filter_label = report_data['employee_filter_label']
        records = report_data['records']
        summaries = report_data['employee_summaries']
        summary_stats = report_data['summary']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Page Setup
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_options.horizontalCentered = True

        # Styles definition
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        card_top_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='none')
        )
        card_bot_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='none'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=9)
        section_font = Font(color="1E293B", bold=True, size=11)
        even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        kpi_bg_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        summary_footer_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        in_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        out_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

        # 1. Logo Embedding
        sp_logo_path = get_logo_path('logo_spinfotech.jpeg')
        vkt_logo_path = get_logo_path('logo_vkt.jpg')

        if sp_logo_path and os.path.exists(sp_logo_path):
            try:
                img_sp = OpenPyXLImage(sp_logo_path)
                img_sp.width = 48
                img_sp.height = 45
                ws.add_image(img_sp, 'A1')
            except Exception as e:
                logger.warning(f"Could not embed S.P.Infotech logo: {e}")

        if vkt_logo_path and os.path.exists(vkt_logo_path):
            try:
                img_vkt = OpenPyXLImage(vkt_logo_path)
                img_vkt.width = 160
                img_vkt.height = 38
                ws.add_image(img_vkt, 'G1')
            except Exception as e:
                logger.warning(f"Could not embed VKT logo: {e}")

        # Row Heights for Header
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 18
        ws.row_dimensions[5].height = 8

        # Header Hierarchy (Columns B to F merged for optimal presentation)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
        c1 = ws.cell(row=1, column=2, value="S.P. INFOTECH")
        c1.font = Font(size=13, bold=True, color="1E293B")
        c1.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)
        c2 = ws.cell(row=2, column=2, value="V.K. TOURS & TRAVELS — ENTERPRISE ATTENDANCE")
        c2.font = Font(size=9, bold=True, color="475569")
        c2.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)
        c3 = ws.cell(row=3, column=2, value=f"ATTENDANCE REPORT  |  DATE RANGE: {display_date_range}")
        c3.font = Font(size=10, bold=True, color="1D4ED8")
        c3.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)
        c4 = ws.cell(row=4, column=2, value=f"Employees: {emp_filter_label}")
        c4.font = Font(size=9, bold=False, color="64748B")
        c4.alignment = Alignment(horizontal="left", vertical="center")

        # 2. Executive Summary KPI Cards (Rows 6 and 7) - 3 Cards (No inferred absence)
        ws.row_dimensions[6].height = 16
        ws.row_dimensions[7].height = 24

        kpis = [
            ({'start': 1, 'end': 2}, "TOTAL ENROLLED", str(summary_stats['enrolled_count']), "64748B", "1E293B"),
            ({'start': 3, 'end': 5}, "PRESENT STAFF", str(summary_stats['present_count']), "16A34A", "16A34A"),
            ({'start': 6, 'end': 7}, "TOTAL PUNCH EVENTS", str(summary_stats['total_events']), "1D4ED8", "1D4ED8"),
        ]

        for span, label, val, label_color, val_color in kpis:
            s_col, e_col = span['start'], span['end']
            if s_col != e_col:
                ws.merge_cells(start_row=6, start_column=s_col, end_row=6, end_column=e_col)
                ws.merge_cells(start_row=7, start_column=s_col, end_row=7, end_column=e_col)

            lbl_cell = ws.cell(row=6, column=s_col, value=label)
            lbl_cell.font = Font(size=8, bold=True, color=label_color)
            lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
            lbl_cell.fill = kpi_bg_fill

            val_cell = ws.cell(row=7, column=s_col, value=val)
            val_cell.font = Font(size=14, bold=True, color=val_color)
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            val_cell.fill = kpi_bg_fill

            # Apply borders to merged range
            for col in range(s_col, e_col + 1):
                c_top = ws.cell(row=6, column=col)
                c_top.fill = kpi_bg_fill
                c_top.border = card_top_border
                c_bot = ws.cell(row=7, column=col)
                c_bot.fill = kpi_bg_fill
                c_bot.border = card_bot_border

        ws.row_dimensions[8].height = 10  # Spacer row

        # 3. Section 1: Employee Presence Summary
        ws.row_dimensions[9].height = 22
        ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=7)
        s1_title = ws.cell(row=9, column=1, value="1. Employee Presence Summary")
        s1_title.font = section_font
        s1_title.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[10].height = 24
        # Columns across 7 columns: Emp ID (A), Name (B-C merged), Dept (D), Days Present (E), Punches (F), Dates (G)
        sum_headers = ['Emp ID', 'Employee Name', 'Department', 'Days Present', 'Total Punches', 'Attendance Dates']
        # We can use col 1: Emp ID, 2: Name, 3: Dept, 4: Days Present, 5: Total Punches, 6-7: Attendance Dates
        ws.cell(row=10, column=1, value='Emp ID').alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=10, column=2, value='Employee Name').alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=10, column=3, value='Department').alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=10, column=4, value='Days Present').alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=10, column=5, value='Total Punches').alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=10, start_column=6, end_row=10, end_column=7)
        ws.cell(row=10, column=6, value='Attendance Dates').alignment = Alignment(horizontal="left", vertical="center")

        for col_num in range(1, 8):
            cell = ws.cell(row=10, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        curr_row = 11
        for s in summaries:
            row_fill = even_fill if curr_row % 2 == 0 else odd_fill
            dates_str = ", ".join(s['dates_present']) if s['dates_present'] else "No punches recorded"

            ws.cell(row=curr_row, column=1, value=s['emp_id']).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=curr_row, column=2, value=s['emp_name']).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=curr_row, column=3, value=s['department']).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=curr_row, column=4, value=s['present_days']).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=curr_row, column=5, value=s['total_punches']).alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=curr_row, start_column=6, end_row=curr_row, end_column=7)
            ws.cell(row=curr_row, column=6, value=dates_str).alignment = Alignment(horizontal="left", vertical="center")

            for col_idx in range(1, 8):
                c = ws.cell(row=curr_row, column=col_idx)
                c.fill = row_fill
                c.border = thin_border
                c.font = Font(size=9, color="0F172A" if s['has_attendance'] else "64748B")

            ws.row_dimensions[curr_row].height = 20
            curr_row += 1

        # Spacer between sections
        ws.row_dimensions[curr_row].height = 12
        curr_row += 1

        # 4. Section 2: Detailed Punch Logs
        ws.row_dimensions[curr_row].height = 22
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
        s2_title = ws.cell(row=curr_row, column=1, value="2. Detailed Punch Logs")
        s2_title.font = section_font
        s2_title.alignment = Alignment(horizontal="left", vertical="center")
        curr_row += 1

        # Freeze panes at the detailed punch logs
        log_header_row_idx = curr_row
        ws.freeze_panes = f'A{log_header_row_idx + 1}'

        ws.row_dimensions[curr_row].height = 24
        log_headers = ['Emp ID', 'Employee Name', 'Department', 'Date', 'Time', 'Punch Type', 'Timestamp']
        for col_num, h_text in enumerate(log_headers, 1):
            cell = ws.cell(row=curr_row, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_num in (1, 4, 5, 6, 7) else "left", vertical="center")
        curr_row += 1

        if not records:
            ws.row_dimensions[curr_row].height = 20
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
            empty_cell = ws.cell(row=curr_row, column=1, value="No punch events recorded in this period")
            empty_cell.alignment = Alignment(horizontal="center", vertical="center")
            empty_cell.font = Font(size=9, italic=True, color="64748B")
            for col_idx in range(1, 8):
                ws.cell(row=curr_row, column=col_idx).fill = even_fill
                ws.cell(row=curr_row, column=col_idx).border = thin_border
            curr_row += 1
        else:
            for r in records:
                row_fill = even_fill if curr_row % 2 == 0 else odd_fill
                punch_type_val = r['type']

                row_values = [
                    r['emp_id'],
                    r['emp_name'],
                    r['department'],
                    r['date'],
                    r['time'],
                    r['type'],
                    r['timestamp'],
                ]

                for col_idx, val in enumerate(row_values, 1):
                    c = ws.cell(row=curr_row, column=col_idx, value=val)
                    c.fill = row_fill
                    c.border = thin_border
                    c.font = Font(size=9, color="0F172A")

                    # Highlight Punch Type cell
                    if col_idx == 6:
                        if punch_type_val == 'IN':
                            c.fill = in_fill
                            c.font = Font(size=9, bold=True, color="15803D")
                        elif punch_type_val == 'OUT':
                            c.fill = out_fill
                            c.font = Font(size=9, bold=True, color="B91C1C")

                    if col_idx in (1, 4, 5, 6, 7):
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center")

                ws.row_dimensions[curr_row].height = 20
                curr_row += 1

        # 5. Totals / Summary Footer Row
        ws.row_dimensions[curr_row].height = 22
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
        foot_label = ws.cell(row=curr_row, column=1, value=f"Total Events: {summary_stats['total_events']}  |  Present Staff: {summary_stats['present_count']} / {summary_stats['enrolled_count']}")
        foot_label.font = Font(size=9, bold=True, color="475569")
        foot_label.alignment = Alignment(horizontal="left", vertical="center")
        foot_label.fill = summary_footer_fill

        for col in range(1, 6):
            ws.cell(row=curr_row, column=col).fill = summary_footer_fill
            ws.cell(row=curr_row, column=col).border = thin_border

        foot_stat = ws.cell(row=curr_row, column=6, value="TOTAL")
        foot_stat.font = Font(size=9, bold=True, color="1E293B")
        foot_stat.alignment = Alignment(horizontal="center", vertical="center")
        foot_stat.fill = summary_footer_fill
        foot_stat.border = thin_border

        foot_cnt = ws.cell(row=curr_row, column=7, value=f"{summary_stats['total_events']} punch events")
        foot_cnt.font = Font(size=9, bold=True, color="1D4ED8")
        foot_cnt.alignment = Alignment(horizontal="center", vertical="center")
        foot_cnt.fill = summary_footer_fill
        foot_cnt.border = thin_border

        # Set Column Widths
        col_widths = {
            'A': 16, # Emp ID
            'B': 24, # Employee Name
            'C': 20, # Department
            'D': 16, # Days Present / Date
            'E': 14, # Total Punches / Time
            'F': 16, # Status / Punch Type
            'G': 32, # Attendance Dates / Timestamp
        }
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Attendance_Report_{file_date_suffix}.xlsx"'
        wb.save(response)
        return response


class ExportPdfReportView(APIView):
    """
    Generates structured corporate PDF report using ReportLab.
    Matches the exact canonical attendance dataset, date range representation, and corporate styling.
    Technical columns (Confidence Score and Record UUID) are excluded from the presentation.
    """
    def get(self, request):
        if not is_authenticated_request(request):
            return Response({'error': 'Authentication credentials were not provided or are invalid'}, status=status.HTTP_401_UNAUTHORIZED)

        start_date = request.query_params.get('start_date') or request.query_params.get('date')
        end_date = request.query_params.get('end_date') or request.query_params.get('date')
        emp_ids = request.query_params.get('emp_ids') or request.query_params.get('emp_id')

        report_data = get_canonical_attendance_data(start_date_str=start_date, end_date_str=end_date, emp_ids=emp_ids)
        display_date_range = report_data['display_date_range']
        file_date_suffix = report_data['file_date_suffix']
        emp_filter_label = report_data['employee_filter_label']
        records = report_data['records']
        summaries = report_data['employee_summaries']
        summary_stats = report_data['summary']

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=20,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('RTitle', fontName='Helvetica-Bold', fontSize=14, leading=17, textColor=rl_colors.HexColor('#1E293B'))
        sub_style = ParagraphStyle('RSub', fontName='Helvetica-Bold', fontSize=9, leading=12, textColor=rl_colors.HexColor('#475569'))
        meta_style = ParagraphStyle('RMeta', fontName='Helvetica-Bold', fontSize=10, leading=13, textColor=rl_colors.HexColor('#1D4ED8'))
        filter_style = ParagraphStyle('RFilter', fontName='Helvetica', fontSize=9, leading=12, textColor=rl_colors.HexColor('#64748B'))
        section_style = ParagraphStyle('RSect', fontName='Helvetica-Bold', fontSize=11, leading=14, textColor=rl_colors.HexColor('#1E293B'), spaceBefore=8, spaceAfter=4)
        th_style = ParagraphStyle('RTH', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=rl_colors.white, alignment=1)
        td_style = ParagraphStyle('RTD', fontName='Helvetica', fontSize=8, leading=10, textColor=rl_colors.HexColor('#0F172A'))
        td_center = ParagraphStyle('RTDC', fontName='Helvetica', fontSize=8, leading=10, textColor=rl_colors.HexColor('#0F172A'), alignment=1)
        td_in = ParagraphStyle('RTDIN', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=rl_colors.HexColor('#15803D'), alignment=1)
        td_out = ParagraphStyle('RTDOUT', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=rl_colors.HexColor('#B91C1C'), alignment=1)

        elements = []

        # 1. Header with Logos
        sp_logo_path = get_logo_path('logo_spinfotech.jpeg')
        vkt_logo_path = get_logo_path('logo_vkt.jpg')

        header_cells = []
        if sp_logo_path and os.path.exists(sp_logo_path):
            img_sp = RLImage(sp_logo_path, width=44, height=40)
            header_cells.append(img_sp)
        else:
            header_cells.append('')

        header_text = [
            Paragraph("S.P. INFOTECH", title_style),
            Paragraph("V.K. TOURS & TRAVELS — ENTERPRISE ATTENDANCE", sub_style),
            Paragraph(f"ATTENDANCE REPORT  |  DATE RANGE: {display_date_range}", meta_style),
            Paragraph(f"Employees: {emp_filter_label}", filter_style),
        ]
        header_cells.append(header_text)

        if vkt_logo_path and os.path.exists(vkt_logo_path):
            img_vkt = RLImage(vkt_logo_path, width=120, height=28)
            header_cells.append(img_vkt)
        else:
            header_cells.append('')

        header_table = Table([header_cells], colWidths=[50, 600, 130])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 8))

        # 2. Executive Summary Metrics Cards - 3 KPI boxes (No inferred absence)
        summary_data = [
            [
                Paragraph("TOTAL ENROLLED", ParagraphStyle('S1', fontName='Helvetica-Bold', fontSize=7, textColor=rl_colors.HexColor('#64748B'), alignment=1)),
                Paragraph("PRESENT STAFF", ParagraphStyle('S2', fontName='Helvetica-Bold', fontSize=7, textColor=rl_colors.HexColor('#16A34A'), alignment=1)),
                Paragraph("TOTAL PUNCH EVENTS", ParagraphStyle('S4', fontName='Helvetica-Bold', fontSize=7, textColor=rl_colors.HexColor('#1D4ED8'), alignment=1)),
            ],
            [
                Paragraph(str(summary_stats['enrolled_count']), ParagraphStyle('V1', fontName='Helvetica-Bold', fontSize=14, textColor=rl_colors.HexColor('#1E293B'), alignment=1)),
                Paragraph(str(summary_stats['present_count']), ParagraphStyle('V2', fontName='Helvetica-Bold', fontSize=14, textColor=rl_colors.HexColor('#16A34A'), alignment=1)),
                Paragraph(str(summary_stats['total_events']), ParagraphStyle('V4', fontName='Helvetica-Bold', fontSize=14, textColor=rl_colors.HexColor('#1D4ED8'), alignment=1)),
            ]
        ]
        sum_table = Table(summary_data, colWidths=[260, 260, 260])
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), rl_colors.HexColor('#F8FAFC')),
            ('BOX', (0, 0), (-1, -1), 1, rl_colors.HexColor('#E2E8F0')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E2E8F0')),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(sum_table)
        elements.append(Spacer(1, 10))

        # 3. Employee Presence Summary Table
        elements.append(Paragraph("1. Employee Presence Summary", section_style))
        emp_sum_headers = [
            Paragraph("Emp ID", th_style),
            Paragraph("Employee Name", th_style),
            Paragraph("Department", th_style),
            Paragraph("Days Present", th_style),
            Paragraph("Total Punches", th_style),
            Paragraph("Attendance Dates", th_style),
        ]
        emp_sum_rows = [emp_sum_headers]

        for s in summaries:
            dates_str = ", ".join(s['dates_present']) if s['dates_present'] else "No punches recorded"
            emp_sum_rows.append([
                Paragraph(s['emp_id'], td_center),
                Paragraph(s['emp_name'], td_style),
                Paragraph(s['department'], td_style),
                Paragraph(str(s['present_days']), td_center),
                Paragraph(str(s['total_punches']), td_center),
                Paragraph(dates_str, td_style),
            ])

        emp_sum_table = Table(emp_sum_rows, colWidths=[70, 150, 130, 85, 85, 260])
        emp_sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1E293B')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#F8FAFC')]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(emp_sum_table)
        elements.append(Spacer(1, 10))

        # 4. Detailed Attendance Punch Log (Technical columns Confidence Score & UUID excluded)
        elements.append(Paragraph("2. Detailed Punch Logs", section_style))
        log_headers = [
            Paragraph("Emp ID", th_style),
            Paragraph("Employee Name", th_style),
            Paragraph("Department", th_style),
            Paragraph("Date", th_style),
            Paragraph("Time", th_style),
            Paragraph("Type", th_style),
            Paragraph("Timestamp", th_style),
        ]
        log_rows = [log_headers]

        if not records:
            log_rows.append([
                Paragraph("-", td_center),
                Paragraph("No punch events recorded in this period", td_style),
                Paragraph("-", td_center),
                Paragraph("-", td_center),
                Paragraph("-", td_center),
                Paragraph("-", td_center),
                Paragraph("-", td_center),
            ])
        else:
            for r in records:
                p_style = td_in if r['type'] == 'IN' else td_out
                log_rows.append([
                    Paragraph(r['emp_id'], td_center),
                    Paragraph(r['emp_name'], td_style),
                    Paragraph(r['department'], td_style),
                    Paragraph(r['date'], td_center),
                    Paragraph(r['time'], td_center),
                    Paragraph(r['type'], p_style),
                    Paragraph(r['timestamp'], td_center),
                ])

        log_table = Table(log_rows, colWidths=[65, 140, 120, 90, 80, 85, 200])
        log_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1E293B')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#F8FAFC')]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(log_table)

        doc.build(elements)
        pdf_content = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Attendance_Report_{file_date_suffix}.pdf"'
        response.write(pdf_content)
        return response



