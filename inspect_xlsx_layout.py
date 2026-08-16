import openpyxl

wb = openpyxl.load_workbook('test_attendance_report_v2.xlsx')
ws = wb.active

print("Sheet Title:", ws.title)
print("Freeze Panes:", ws.freeze_panes)
print("Page Setup Orientation:", ws.page_setup.orientation)
print("Fit To Page Enabled:", ws.sheet_properties.pageSetUpPr.fitToPage)
print("Print Title Rows:", ws.print_title_rows)

print("\n--- Header Structure ---")
for r in range(1, 6):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
    print(f"Row {r}: {row_vals}")

print("\n--- Data Sample ---")
for r in range(6, 10):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
    print(f"Row {r}: {row_vals}")

print("\n--- Column Widths ---")
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    print(f"Column {col_letter}: {ws.column_dimensions[col_letter].width}")

print("\n--- Embedded Images ---")
print("Images count:", len(ws._images))
for idx, img in enumerate(ws._images, 1):
    print(f"Image {idx}: width={img.width}, height={img.height}")
